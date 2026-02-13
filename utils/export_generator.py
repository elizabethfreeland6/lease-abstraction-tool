"""
Export Generation Module
Generates Yardi-compatible Excel files and reference documents
"""

import os
from datetime import datetime
from typing import List, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def generate_yardi_excel(extracted_data: List[Dict], output_dir: str = "exports") -> str:
    """
    Generate Yardi-compatible Excel file for automatic import
    
    Args:
        extracted_data: List of dictionaries containing extracted lease data
        output_dir: Directory to save the output file
        
    Returns:
        Path to the generated Excel file
    """
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Prepare data for Yardi format
    yardi_data = []
    
    for doc in extracted_data:
        data = doc['data']
        
        # Map extracted data to Yardi columns
        yardi_row = {
            'TenantName': data.get('tenant_name', ''),
            'TenantEmail': data.get('tenant_email', ''),
            'TenantPhone': data.get('tenant_phone', ''),
            'PropertyAddress': data.get('property_address', ''),
            'UnitNumber': data.get('unit_number', ''),
            'PropertyType': data.get('property_type', ''),
            'SquareFootage': data.get('square_footage', 0),
            'LeaseNumber': data.get('lease_number', ''),
            'LeaseStartDate': data.get('lease_start_date', ''),
            'LeaseEndDate': data.get('lease_end_date', ''),
            'LeaseTermMonths': data.get('lease_term_months', 0),
            'LeaseType': data.get('lease_type', ''),
            'MonthlyRent': data.get('monthly_rent', 0),
            'SecurityDeposit': data.get('security_deposit', 0),
            'PetDeposit': data.get('pet_deposit', 0),
            'PaymentDueDate': data.get('payment_due_date', 1),
            'LateFeeType': data.get('late_fee_type', ''),
            'LateFeePercentage': data.get('late_fee_percentage', 0),
            'LateFeeAmount': data.get('late_fee_flat_amount', 0),
            'LateFeeGracePeriod': data.get('late_fee_grace_period', 0),
            'ParkingSpaces': data.get('parking_spaces', 0),
            'PetAllowed': 'Yes' if data.get('pet_allowed', False) else 'No',
            'PetType': data.get('pet_type', ''),
            'UtilitiesIncluded': data.get('utilities_included', ''),
            'SourceFile': doc.get('filename', '')
        }
        
        yardi_data.append(yardi_row)
    
    # Create DataFrame
    df = pd.DataFrame(yardi_data)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"yardi_import_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    # Create Excel workbook with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Yardi Import"
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    
    # Adjust column widths
    column_widths = {
        'A': 25,  # TenantName
        'B': 30,  # TenantEmail
        'C': 15,  # TenantPhone
        'D': 40,  # PropertyAddress
        'E': 12,  # UnitNumber
        'F': 15,  # PropertyType
        'G': 15,  # SquareFootage
        'H': 15,  # LeaseNumber
        'I': 15,  # LeaseStartDate
        'J': 15,  # LeaseEndDate
        'K': 15,  # LeaseTermMonths
        'L': 15,  # LeaseType
        'M': 12,  # MonthlyRent
        'N': 15,  # SecurityDeposit
        'O': 12,  # PetDeposit
        'P': 15,  # PaymentDueDate
        'Q': 15,  # LateFeeAmount
        'R': 18,  # LateFeeGracePeriod
        'S': 15,  # ParkingSpaces
        'T': 12,  # PetAllowed
        'U': 15,  # PetType
        'V': 25,  # EmergencyContactName
        'W': 20,  # EmergencyContactPhone
        'X': 30,  # UtilitiesIncluded
        'Y': 30   # SourceFile
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(filepath)
    
    return filepath

def generate_reference_document(extracted_data: List[Dict], output_dir: str = "exports") -> str:
    """
    Generate comprehensive reference document with all extracted data
    
    Args:
        extracted_data: List of dictionaries containing extracted lease data
        output_dir: Directory to save the output file
        
    Returns:
        Path to the generated reference document
    """
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"lease_reference_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    # Create Excel workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define styles
    section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(color="FFFFFF", bold=True, size=12)
    
    field_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    field_font = Font(bold=True, size=10)
    
    value_font = Font(size=10)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create a sheet for each document
    for idx, doc in enumerate(extracted_data, 1):
        data = doc['data']
        filename_short = doc['filename'][:25]  # Truncate for sheet name
        
        ws = wb.create_sheet(title=f"Lease_{idx}")
        
        current_row = 1
        
        # Document header
        ws.merge_cells(f'A{current_row}:B{current_row}')
        cell = ws.cell(row=current_row, column=1, value=f"LEASE ABSTRACTION - {doc['filename']}")
        cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
        
        # Extraction info
        ws.merge_cells(f'A{current_row}:B{current_row}')
        cell = ws.cell(row=current_row, column=1, value=f"Extracted: {doc.get('extracted_at', 'N/A')}")
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(horizontal="center")
        current_row += 2
        
        # Tenant Information Section
        current_row = add_section(ws, current_row, "TENANT INFORMATION", [
            ("Tenant Name", data.get('tenant_name', '')),
            ("Tenant Email", data.get('tenant_email', '')),
            ("Tenant Phone", data.get('tenant_phone', ''))
        ], section_fill, section_font, field_fill, field_font, value_font, border)
        
        # Note about emergency contacts
        ws.merge_cells(f'A{current_row}:B{current_row}')
        note_cell = ws.cell(row=current_row, column=1, value="Note: Emergency contact info typically provided in welcome letter")
        note_cell.font = Font(italic=True, size=9, color="666666")
        note_cell.alignment = Alignment(horizontal="left")
        current_row += 1
        
        # Property Information Section
        current_row = add_section(ws, current_row, "PROPERTY INFORMATION", [
            ("Property Address", data.get('property_address', '')),
            ("Unit Number", data.get('unit_number', '')),
            ("Property Type", data.get('property_type', '')),
            ("Square Footage", data.get('square_footage', 0))
        ], section_fill, section_font, field_fill, field_font, value_font, border)
        
        # Lease Terms Section
        current_row = add_section(ws, current_row, "LEASE TERMS", [
            ("Lease Number", data.get('lease_number', '')),
            ("Lease Start Date", data.get('lease_start_date', '')),
            ("Lease End Date", data.get('lease_end_date', '')),
            ("Lease Term (Months)", data.get('lease_term_months', 0)),
            ("Lease Type", data.get('lease_type', ''))
        ], section_fill, section_font, field_fill, field_font, value_font, border)
        
        # Financial Terms Section
        late_fee_display = ""
        if data.get('late_fee_type') == 'percentage':
            late_fee_display = f"{data.get('late_fee_percentage', 0)}% of payment"
        elif data.get('late_fee_type') == 'flat_amount':
            late_fee_display = f"${data.get('late_fee_flat_amount', 0):.2f}"
        else:
            late_fee_display = "Not specified"
        
        current_row = add_section(ws, current_row, "FINANCIAL TERMS", [
            ("Monthly Rent", f"${data.get('monthly_rent', 0):.2f}"),
            ("Security Deposit", f"${data.get('security_deposit', 0):.2f}"),
            ("Pet Deposit", f"${data.get('pet_deposit', 0):.2f}"),
            ("Payment Due Date", f"Day {int(data.get('payment_due_date', 1))} of month"),
            ("Late Fee Type", data.get('late_fee_type', 'Not specified')),
            ("Late Fee", late_fee_display),
            ("Late Fee Grace Period", f"{int(data.get('late_fee_grace_period', 0))} days")
        ], section_fill, section_font, field_fill, field_font, value_font, border)
        
        # Additional Terms Section
        current_row = add_section(ws, current_row, "ADDITIONAL TERMS", [
            ("Parking Spaces", int(data.get('parking_spaces', 0))),
            ("Pet Allowed", "Yes" if data.get('pet_allowed', False) else "No"),
            ("Pet Type", data.get('pet_type', '')),
            ("Utilities Included", data.get('utilities_included', '')),
            ("Renewal Options", data.get('renewal_options', '')),
            ("Early Termination", data.get('early_termination_clause', '')),
            ("Maintenance Responsibilities", data.get('maintenance_responsibilities', ''))
        ], section_fill, section_font, field_fill, field_font, value_font, border)
        
        # Confidence Score
        confidence = data.get('confidence_score', 0.5)
        confidence_level = "High" if confidence >= 0.8 else "Medium" if confidence >= 0.5 else "Low"
        confidence_color = "28A745" if confidence >= 0.8 else "FFC107" if confidence >= 0.5 else "DC3545"
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:B{current_row}')
        cell = ws.cell(row=current_row, column=1, value=f"Extraction Confidence: {confidence_level} ({confidence:.2%})")
        cell.fill = PatternFill(start_color=confidence_color, end_color=confidence_color, fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
    
    # Create summary sheet
    create_summary_sheet(wb, extracted_data)
    
    # Save workbook
    wb.save(filepath)
    
    return filepath

def add_section(ws, start_row, section_title, fields, section_fill, section_font, 
                field_fill, field_font, value_font, border):
    """
    Add a section to the worksheet
    
    Returns the next available row number
    """
    # Section header
    ws.merge_cells(f'A{start_row}:B{start_row}')
    cell = ws.cell(row=start_row, column=1, value=section_title)
    cell.fill = section_fill
    cell.font = section_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
    start_row += 1
    
    # Fields
    for field_name, field_value in fields:
        # Field name
        cell = ws.cell(row=start_row, column=1, value=field_name)
        cell.fill = field_fill
        cell.font = field_font
        cell.border = border
        cell.alignment = Alignment(vertical="center")
        
        # Field value
        cell = ws.cell(row=start_row, column=2, value=field_value)
        cell.font = value_font
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        
        start_row += 1
    
    start_row += 1  # Add spacing after section
    return start_row

def create_summary_sheet(wb, extracted_data):
    """
    Create a summary sheet with overview of all leases
    """
    ws = wb.create_sheet(title="Summary", index=0)
    
    # Header
    ws.merge_cells('A1:H1')
    cell = ws.cell(row=1, column=1, value="LEASE ABSTRACTION SUMMARY")
    cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Column headers
    headers = ["#", "Tenant Name", "Property Address", "Unit", "Start Date", "End Date", "Monthly Rent", "Confidence"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for idx, doc in enumerate(extracted_data, 1):
        data = doc['data']
        row = idx + 3
        
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=data.get('tenant_name', ''))
        ws.cell(row=row, column=3, value=data.get('property_address', ''))
        ws.cell(row=row, column=4, value=data.get('unit_number', ''))
        ws.cell(row=row, column=5, value=data.get('lease_start_date', ''))
        ws.cell(row=row, column=6, value=data.get('lease_end_date', ''))
        ws.cell(row=row, column=7, value=f"${data.get('monthly_rent', 0):.2f}")
        
        confidence = data.get('confidence_score', 0.5)
        confidence_level = "High" if confidence >= 0.8 else "Medium" if confidence >= 0.5 else "Low"
        ws.cell(row=row, column=8, value=confidence_level)
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    
    # Freeze panes
    ws.freeze_panes = 'A4'
